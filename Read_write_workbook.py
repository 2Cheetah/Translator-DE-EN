from openpyxl import load_workbook


def read_from_xlsx(filename, start_row, end_row, column):
    wb = load_workbook(filename=filename)
    text_to_translate = []

    ws = wb.active

    for i in range(start_row, end_row, 1):
        text_to_translate.append(ws[column + str(i)].value)

    return text_to_translate


def write_to_xlsx(filename, start_row, column, translated_text):
    wb = load_workbook(filename=filename)
    ws = wb.active
    i = start_row

    for translated_sentence in translated_text:
        ws[column + str(i)].value = translated_sentence
        i += 1

    wb.save(filename=filename)
